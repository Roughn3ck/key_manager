"""
Key Manager - Secure Crypto Key Storage
CLI interface for managing encrypted cryptocurrency keys and addresses.
"""
import json
import os
import sys
import csv
import base64
from pathlib import Path

# Fix Windows console encoding for Unicode characters (✓, ✗, etc.)
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
from typing import Dict, List, Optional, Any

import click
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn

from crypto_engine import CryptoEngine, generate_secure_password


console = Console()
crypto = CryptoEngine()

# Constants
DATA_FILE = "key_vault.encrypted"
SESSION_FILE = ".key_manager_session"
CONFIG_DIR = Path.home() / ".key_manager"


def get_data_dir() -> Path:
    """Determine the data directory based on how the app is running.
    
    If running as a frozen EXE (portable/USB mode), use the EXE directory
    so the vault is stored alongside the executable for true portability.
    Otherwise, use the user's home directory (~/.key_manager).
    """
    if getattr(sys, 'frozen', False):
        return Path(os.path.dirname(sys.executable))
    else:
        return CONFIG_DIR


class KeyManager:
    """Main application for managing encrypted keys."""
    
    def __init__(self, data_dir: Optional[Path] = None) -> None:
        self.data_dir = data_dir or get_data_dir()
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.data_file = self.data_dir / DATA_FILE
        self.session_file = self.data_dir / SESSION_FILE
        
        # Start with an empty vault — the user creates pools and accounts
        # as needed.  Previously this was pre-populated with development
        # pools/accounts (Genesis, SafetyNet, etc.) which were specific to
        # the original developer and not appropriate for a general release.
        self.address_db = {
            "version": "1.0",
            "pools": {},
            "accounts": {},
            "mnemonics": {},
            "private_keys": {}
        }
    
    def load_encrypted_data(self, password: str) -> bool:
        """Load encrypted data from file, replacing current address_db."""
        if not self.data_file.exists():
            return False
        
        try:
            with open(self.data_file, 'r') as f:
                encrypted_json = f.read()
            
            decrypted_data = crypto.decrypt_json(encrypted_json, password)
            
            # Replace entire address_db with loaded data
            self.address_db = decrypted_data
            return True
            
        except Exception as e:
            console.print(f"[red]Error loading data: {e}[/red]")
            return False
    
    def save_encrypted_data(self, password: str) -> bool:
        """Save encrypted data to file."""
        try:
            encrypted_json = crypto.encrypt_json(self.address_db, password)
            with open(self.data_file, 'w') as f:
                f.write(encrypted_json)
            return True
        except Exception as e:
            console.print(f"[red]Error saving data: {e}[/red]")
            return False
    
    def create_session(self, password: str) -> None:
        """Store password in session file for use by subsequent commands."""
        encoded = base64.b64encode(password.encode()).decode()
        with open(self.session_file, 'w') as f:
            f.write(encoded)
    
    def get_session_password(self) -> Optional[str]:
        """Retrieve password from session file."""
        if not self.session_file.exists():
            return None
        try:
            with open(self.session_file, 'r') as f:
                encoded = f.read().strip()
            if not encoded:
                return None
            return base64.b64decode(encoded.encode()).decode()
        except Exception:
            return None
    
    def end_session(self) -> None:
        """Remove session file."""
        if self.session_file.exists():
            self.session_file.unlink()
    
    def is_session_active(self) -> bool:
        """Check if a session is active."""
        return self.session_file.exists()
    
    def initialize_vault(self, password: str) -> bool:
        """Initialize a new encrypted vault."""
        if self.data_file.exists():
            console.print("[yellow]Vault already exists. Use 'unlock' instead.[/yellow]")
            return False
        
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            transient=True,
        ) as progress:
            progress.add_task(description="Initializing secure vault...", total=None)
            success = self.save_encrypted_data(password)
        
        if success:
            console.print("[green]✓ Secure vault initialized successfully[/green]")
            return True
        else:
            console.print("[red]Failed to initialize vault[/red]")
            return False
    
    def add_mnemonic(self, account_name: str, mnemonic: str, password: str) -> bool:
        """Add or update a mnemonic for an account."""
        self.address_db.setdefault("mnemonics", {})[account_name] = mnemonic
        return self.save_encrypted_data(password)
    
    def show_mnemonic(self, account_name: str) -> Optional[str]:
        """Get mnemonic for an account, or None if not found."""
        return self.address_db.get("mnemonics", {}).get(account_name)
    
    def add_account(self, account_name: str, pool: Optional[str] = None) -> bool:
        """Create a new account entry, optionally in a pool."""
        if account_name in self.address_db.get("accounts", {}):
            return False
        self.address_db.setdefault("accounts", {})[account_name] = {"addresses": [], "notes": ""}
        if pool and pool in self.address_db.get("pools", {}):
            self.address_db["pools"][pool].setdefault("accounts", []).append(account_name)
        return True
    
    def add_private_key(self, account_name: str, private_key: str, password: str,
                        chain: str = "") -> bool:
        """Add a chain-specific private key to an account.

        Stores keys as a list of {chain, key} dicts per account, supporting
        multiple chain-specific keys. Backward-compatible: if existing data
        is a plain string, it is migrated to a list on first append.
        """
        pk_store = self.address_db.setdefault("private_keys", {})
        existing = pk_store.get(account_name)

        # Migrate legacy single-string format to list format
        if existing is None:
            pk_store[account_name] = []
        elif isinstance(existing, str):
            pk_store[account_name] = [{"chain": "", "key": existing}]

        pk_store[account_name].append({"chain": chain, "key": private_key})
        return self.save_encrypted_data(password)

    def show_private_key(self, account_name: str) -> List[Dict[str, str]]:
        """Get private keys for an account, or empty list if not found.

        Returns a list of {chain, key} dicts. Backward-compatible: if legacy
        single-string format is found, returns it as a single-item list.
        """
        existing = self.address_db.get("private_keys", {}).get(account_name)
        if existing is None:
            return []
        if isinstance(existing, str):
            return [{"chain": "", "key": existing}]
        return existing
    
    def change_password(self, old_password: str, new_password: str) -> bool:
        """Change the vault password by re-encrypting with new password."""
        if not self.data_file.exists():
            return False
        try:
            with open(self.data_file, 'r') as f:
                encrypted_json = f.read()
            decrypted_data = crypto.decrypt_json(encrypted_json, old_password)
            self.address_db = decrypted_data
            new_encrypted = crypto.encrypt_json(self.address_db, new_password)
            with open(self.data_file, 'w') as f:
                f.write(new_encrypted)
            self.create_session(new_password)
            return True
        except Exception:
            return False
    
    def add_address(self, account: str, coin: str, chain: str, address: str,
                    password: str, notes: str = "") -> bool:
        """Add an address to an account."""
        if account not in self.address_db["accounts"]:
            self.address_db["accounts"][account] = {
                "addresses": [],
                "notes": ""
            }
        
        addr_entry = {
            "coin": coin,
            "chain": chain,
            "address": address
        }
        if notes:
            addr_entry["notes"] = notes
        
        self.address_db["accounts"][account]["addresses"].append(addr_entry)
        return self.save_encrypted_data(password)

    def delete_account(self, account_name: str, password: str) -> bool:
        """Delete an account and all its data (addresses, mnemonic, private keys).

        Also removes the account from any pool it belongs to.

        Args:
            account_name: Name of the account to delete.
            password: Vault password for re-encryption.

        Returns:
            True if the account was found and deleted, False otherwise.
        """
        accounts = self.address_db.get("accounts", {})
        if account_name not in accounts:
            return False

        # Remove the account entry
        del accounts[account_name]

        # Remove from any pool
        for pool_data in self.address_db.get("pools", {}).values():
            if account_name in pool_data.get("accounts", []):
                pool_data["accounts"].remove(account_name)

        # Remove mnemonic if present
        self.address_db.get("mnemonics", {}).pop(account_name, None)

        # Remove private keys if present
        self.address_db.get("private_keys", {}).pop(account_name, None)

        return self.save_encrypted_data(password)

    def delete_address(self, account: str, index: int, password: str) -> bool:
        """Delete an address at the given 0-based index from an account.

        Args:
            account: Account name.
            index: 0-based index into the account's addresses list.
            password: Vault password for re-encryption.

        Returns:
            True if the address was found and deleted, False otherwise.
        """
        accounts = self.address_db.get("accounts", {})
        if account not in accounts:
            return False
        addresses = accounts[account].get("addresses", [])
        if 0 <= index < len(addresses):
            addresses.pop(index)
            return self.save_encrypted_data(password)
        return False
    
    def import_csv(self, csv_path: str, password: str) -> tuple:
        """Import addresses from a CSV file.
        
        Returns:
            Tuple of (added_count, skipped_count, error_messages)
        """
        added = 0
        skipped = 0
        errors = []
        
        try:
            # Use utf-8-sig to strip any BOM (byte order mark) that Excel
            # adds when saving CSV files on Windows.  Without this the first
            # column header becomes '\ufeffAccount' instead of 'Account',
            # causing every row to be skipped.
            with open(csv_path, 'r', newline='', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # Normalise header keys to strip whitespace/BOM remnants
                        normalised_row = {}
                        for k, v in row.items():
                            if k is not None:
                                normalised_row[k.strip()] = v
                        account = normalised_row.get('Account', '').strip()
                        coin = normalised_row.get('Coin', '').strip()
                        chain = normalised_row.get('Chain', '').strip()
                        address = normalised_row.get('Address', '').strip()
                        notes = normalised_row.get('Notes', '').strip()
                        
                        if not account:
                            skipped += 1
                            continue
                        
                        # Skip rows with empty address (account placeholder)
                        if not address:
                            # Still create the account entry
                            if account not in self.address_db["accounts"]:
                                self.address_db["accounts"][account] = {
                                    "addresses": [],
                                    "notes": notes
                                }
                            skipped += 1
                            continue
                        
                        if account not in self.address_db["accounts"]:
                            self.address_db["accounts"][account] = {
                                "addresses": [],
                                "notes": ""
                            }
                        
                        addr_entry = {
                            "coin": coin,
                            "chain": chain,
                            "address": address
                        }
                        if notes:
                            addr_entry["notes"] = notes
                        
                        self.address_db["accounts"][account]["addresses"].append(addr_entry)
                        added += 1
                        
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        skipped += 1
            
            if added > 0:
                self.save_encrypted_data(password)
            
            return (added, skipped, errors)
            
        except FileNotFoundError:
            return (0, 0, [f"CSV file not found: {csv_path}"])
        except Exception as e:
            return (0, 0, [f"Error reading CSV: {str(e)}"])
    
    def import_file(self, file_path: str, password: str) -> tuple:
        """Import addresses from a CSV or Excel (.xlsx/.xls) file.

        Detects file type by extension and delegates to the appropriate
        parser.  Both formats require columns: Account, Address.
        Optional columns: Coin, Chain, Notes.

        Args:
            file_path: Path to the CSV or Excel file.
            password: Vault password for re-encryption.

        Returns:
            Tuple of (added_count, skipped_count, error_messages)
        """
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            return self.import_csv(file_path, password)
        elif ext in ('.xlsx', '.xls'):
            return self.import_excel(file_path, password)
        else:
            return (0, 0, [f"Unsupported file type: {ext} (use .csv, .xlsx, or .xls)"])

    def import_excel(self, excel_path: str, password: str) -> tuple:
        """Import addresses from an Excel (.xlsx/.xls) file.

        Requires openpyxl.  Columns expected: Account, Address.
        Optional: Coin, Chain, Notes.

        Returns:
            Tuple of (added_count, skipped_count, error_messages)
        """
        added = 0
        skipped = 0
        errors = []

        try:
            from openpyxl import load_workbook
        except ImportError:
            return (0, 0, ["openpyxl is required to import Excel files. Install with: pip install openpyxl"])

        try:
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active

            rows = ws.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration:
                return (0, 0, ["Excel file appears to be empty"])

            # Build a lowercase header→index map
            header_map = {}
            for idx, h in enumerate(headers):
                if h is not None:
                    header_map[str(h).strip().lower()] = idx

            # Validate required columns
            if 'account' not in header_map or 'address' not in header_map:
                return (0, 0, ["Excel must have 'Account' and 'Address' columns"])

            for row_num, row in enumerate(rows, start=2):
                try:
                    def get_col(name):
                        idx = header_map.get(name)
                        if idx is None or idx >= len(row):
                            return ''
                        val = row[idx]
                        return str(val).strip() if val is not None else ''

                    account = get_col('account')
                    coin = get_col('coin')
                    chain = get_col('chain')
                    address = get_col('address')
                    notes = get_col('notes')

                    if not account:
                        skipped += 1
                        continue

                    if not address:
                        # Still create the account entry
                        if account not in self.address_db["accounts"]:
                            self.address_db["accounts"][account] = {
                                "addresses": [],
                                "notes": notes
                            }
                        skipped += 1
                        continue

                    if account not in self.address_db["accounts"]:
                        self.address_db["accounts"][account] = {
                            "addresses": [],
                            "notes": ""
                        }

                    addr_entry = {
                        "coin": coin,
                        "chain": chain,
                        "address": address
                    }
                    if notes:
                        addr_entry["notes"] = notes

                    self.address_db["accounts"][account]["addresses"].append(addr_entry)
                    added += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    skipped += 1

            wb.close()

            if added > 0:
                self.save_encrypted_data(password)

            return (added, skipped, errors)

        except FileNotFoundError:
            return (0, 0, [f"Excel file not found: {excel_path}"])
        except Exception as e:
            return (0, 0, [f"Error reading Excel: {str(e)}"])

    def list_accounts(self) -> None:
        """Display all accounts organized by pool."""
        table = Table(title="Accounts by Pool")
        table.add_column("Pool", style="cyan")
        table.add_column("Description", style="yellow")
        table.add_column("Accounts", style="green")
        
        for pool_name, pool_data in self.address_db.get("pools", {}).items():
            accounts = ", ".join(pool_data.get("accounts", []))
            table.add_row(pool_name, pool_data.get("description", ""), accounts)
        
        # Show accounts not in any pool
        all_pool_accounts = set()
        for pool_data in self.address_db.get("pools", {}).values():
            all_pool_accounts.update(pool_data.get("accounts", []))
        
        unassigned = set(self.address_db.get("accounts", {}).keys()) - all_pool_accounts
        if unassigned:
            table.add_row("Unassigned", "Accounts not in a pool", ", ".join(sorted(unassigned)))
        
        console.print(table)
    
    def show_addresses(self, account: Optional[str] = None) -> None:
        """Display addresses for all accounts or a specific account."""
        accounts = self.address_db.get("accounts", {})
        
        if not accounts:
            console.print("[yellow]No accounts with addresses found.[/yellow]")
            console.print("[dim]Use 'add-address' or 'import-csv' to add addresses.[/dim]")
            return
        
        if account:
            if account not in accounts:
                console.print(f"[red]Account '{account}' not found[/red]")
                console.print(f"[dim]Available accounts: {', '.join(sorted(accounts.keys()))}[/dim]")
                return
            self._show_account_addresses(account)
        else:
            for acc_name in sorted(accounts.keys()):
                self._show_account_addresses(acc_name)
    
    def _show_account_addresses(self, account: str) -> None:
        """Display addresses for a specific account."""
        acc_data = self.address_db["accounts"][account]
        notes = acc_data.get("notes", '')
        
        panel_content = f"[bold cyan]{account}[/bold cyan]"
        if notes:
            panel_content += f"\n[yellow]Notes: {notes}[/yellow]"
        
        panel = Panel.fit(panel_content, title=f"Account: {account}")
        console.print(panel)
        
        addresses = acc_data.get("addresses", [])
        if not addresses:
            console.print("[dim]  No addresses stored[/dim]\n")
            return
        
        table = Table(show_header=True, header_style="bold magenta")
        table.add_column("Coin", style="green")
        table.add_column("Chain", style="cyan")
        table.add_column("Address", style="white")
        table.add_column("Notes", style="yellow")
        
        for addr in addresses:
            addr_notes = addr.get("notes", "")
            table.add_row(
                addr.get("coin", ""),
                addr.get("chain", ""),
                addr.get("address", ""),
                addr_notes
            )
        
        console.print(table)
        console.print()


def get_manager_with_session(ctx) -> tuple:
    """Get a KeyManager loaded with session data if available.
    
    Returns:
        Tuple of (manager, password) where password may be None if no session
    """
    manager = ctx.obj.get('manager')
    if not manager:
        return None, None
    
    password = manager.get_session_password()
    if password and manager.data_file.exists():
        manager.load_encrypted_data(password)
    
    return manager, password


@click.group()
@click.pass_context
def cli(ctx):
    """Secure Crypto Key Manager - Store and manage encrypted cryptocurrency keys."""
    ctx.ensure_object(dict)
    manager = KeyManager()
    ctx.obj['manager'] = manager


@cli.command()
@click.option('--password', prompt=True, hide_input=True, confirmation_prompt=True,
              help='Master password for the vault')
@click.pass_context
def init(ctx, password: str):
    """Initialize a new encrypted vault."""
    manager = ctx.obj['manager']
    if manager.initialize_vault(password):
        # Auto-create session after init
        manager.create_session(password)
        console.print("[green]✓ Session started. You can now use other commands without re-entering password.[/green]")


@cli.command()
@click.option('--password', prompt=True, hide_input=True,
              help='Master password to unlock the vault')
@click.pass_context
def unlock(ctx, password: str):
    """Unlock and load an existing vault."""
    manager = ctx.obj['manager']
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        transient=True,
    ) as progress:
        progress.add_task(description="Unlocking vault...", total=None)
        
        if manager.load_encrypted_data(password):
            # Create session for subsequent commands
            manager.create_session(password)
            console.print("[green]✓ Vault unlocked successfully[/green]")
            console.print("[dim]Session active. You can now use other commands without re-entering password.[/dim]")
            
            # Show summary
            accounts = manager.address_db.get("accounts", {})
            addr_count = sum(len(acc.get("addresses", [])) for acc in accounts.values())
            console.print(f"[dim]Loaded {len(accounts)} accounts with {addr_count} addresses[/dim]")
        else:
            console.print("[red]Failed to unlock vault. Wrong password or vault doesn't exist.[/red]")
            sys.exit(1)


@cli.command()
@click.pass_context
def lock(ctx):
    """Lock the vault and end the current session."""
    manager = ctx.obj['manager']
    manager.end_session()
    console.print("[green]✓ Session locked. Password cleared.[/green]")


@cli.command()
@click.pass_context
def accounts(ctx):
    """List all accounts organized by pool."""
    manager, password = get_manager_with_session(ctx)
    
    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)
    
    manager.list_accounts()


@cli.command()
@click.argument('account', required=False)
@click.pass_context
def addresses(ctx, account: Optional[str]):
    """Show addresses for all accounts or a specific account."""
    manager, password = get_manager_with_session(ctx)
    
    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)
    
    manager.show_addresses(account)


@cli.command()
@click.argument('account')
@click.argument('coin')
@click.argument('chain')
@click.argument('address')
@click.option('--notes', default='', help='Optional notes for this address')
@click.pass_context
def add_address(ctx, account: str, coin: str, chain: str, address: str, notes: str):
    """Add an address to an account."""
    manager, password = get_manager_with_session(ctx)
    
    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)
    
    if manager.add_address(account, coin, chain, address, password, notes):
        console.print(f"[green]✓ Address added to {account} successfully[/green]")
    else:
        console.print("[red]Failed to add address[/red]")


@cli.command()
@click.argument('account')
@click.argument('index', type=int)
@click.pass_context
def delete_address(ctx, account: str, index: int):
    """Delete an address from an account by its 0-based index.

    Use 'addresses <account>' to see the index of each address.
    """
    manager, password = get_manager_with_session(ctx)

    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)

    if manager.delete_address(account, index, password):
        console.print(f"[green]✓ Address {index} deleted from {account}[/green]")
    else:
        console.print("[red]Failed to delete address. Check account name and index.[/red]")
        sys.exit(1)


@cli.command()
@click.argument('account')
@click.argument('mnemonic', nargs=-1)
@click.pass_context
def add_mnemonic(ctx, account: str, mnemonic: tuple):
    """Add or update a mnemonic for an account. Use 'show-mnemonic' to view stored mnemonics."""
    manager, password = get_manager_with_session(ctx)
    
    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)
    
    mnemonic_str = ' '.join(mnemonic)
    if len(mnemonic_str.split()) != 24:
        console.print("[red]Mnemonic must be 24 words[/red]")
        sys.exit(1)
    
    if manager.add_mnemonic(account, mnemonic_str, password):
        console.print("[green]✓ Mnemonic added successfully[/green]")
    else:
        console.print("[red]Failed to add mnemonic[/red]")


@cli.command()
@click.argument('account')
@click.option('--pool', default=None, help='Pool to add the account to')
@click.pass_context
def add_account(ctx, account: str, pool: str):
    """Create a new account entry, optionally in a pool."""
    manager, password = get_manager_with_session(ctx)
    
    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)
    
    if manager.add_account(account, pool):
        if password:
            manager.save_encrypted_data(password)
        if pool:
            console.print(f"[green]✓ Account '{account}' created and added to pool '{pool}'[/green]")
        else:
            console.print(f"[green]✓ Account '{account}' created[/green]")
    else:
        console.print(f"[yellow]Account '{account}' already exists[/yellow]")


@cli.command()
@click.argument('account')
@click.pass_context
def show_mnemonic(ctx, account: str):
    """Show the mnemonic for an account. Requires password re-entry for security."""
    manager, password = get_manager_with_session(ctx)
    
    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)
    
    mnemonic = manager.show_mnemonic(account)
    if mnemonic is None:
        console.print(f"[red]No mnemonic found for account '{account}'[/red]")
        sys.exit(1)
    
    # Re-verify password for sensitive operation
    verify = click.prompt("Re-enter password to view mnemonic", hide_input=True)
    if verify != password:
        console.print("[red]Password incorrect[/red]")
        sys.exit(1)
    
    console.print(Panel(
        f"[green]{mnemonic}[/green]",
        title=f"Mnemonic for {account}",
        subtitle="⚠ Copy this securely and clear your terminal afterward"
    ))


@cli.command()
@click.argument('account')
@click.argument('private_key')
@click.option('--chain', default='', help='Chain label for this key (e.g. DASH, BTC, EVM). Leave empty for no chain.')
@click.pass_context
def add_key(ctx, account: str, private_key: str, chain: str):
    """Add a chain-specific private key to an account. Multiple keys per account are supported."""
    manager, password = get_manager_with_session(ctx)
    
    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)
    
    if manager.add_private_key(account, private_key, password, chain):
        chain_label = f" [{chain}]" if chain else ""
        console.print(f"[green]✓ Private key{chain_label} added for {account}[/green]")
    else:
        console.print("[red]Failed to add private key[/red]")


@cli.command()
@click.argument('account')
@click.pass_context
def show_key(ctx, account: str):
    """Show all private keys for an account. Requires password re-entry for security."""
    manager, password = get_manager_with_session(ctx)
    
    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)
    
    keys = manager.show_private_key(account)
    if not keys:
        console.print(f"[red]No private key found for account '{account}'[/red]")
        sys.exit(1)
    
    # Re-verify password for sensitive operation
    verify = click.prompt("Re-enter password to view private keys", hide_input=True)
    if verify != password:
        console.print("[red]Password incorrect[/red]")
        sys.exit(1)
    
    # Display each key with its chain label
    lines = []
    for entry in keys:
        chain = entry.get("chain", "")
        key_val = entry.get("key", "")
        if chain:
            lines.append(f"[cyan][{chain}][/cyan] Key: [green]{key_val}[/green]")
        else:
            lines.append(f"[cyan][No Chain Specified][/cyan] Key: [green]{key_val}[/green]")
    
    panel_content = "\n".join(lines)
    console.print(Panel(
        panel_content,
        title=f"Private Keys for {account}",
        subtitle="⚠ Never share these keys with anyone"
    ))


@cli.command()
@click.pass_context
def change_password(ctx):
    """Change the vault master password."""
    manager = ctx.obj['manager']
    
    if not manager.data_file.exists():
        console.print("[red]No vault found. Initialize first with 'init' command.[/red]")
        sys.exit(1)
    
    old_password = click.prompt("Current password", hide_input=True)
    new_password = click.prompt("New password", hide_input=True, confirmation_prompt=True)
    
    if manager.change_password(old_password, new_password):
        console.print("[green]✓ Password changed successfully[/green]")
        console.print("[dim]New session created with updated password.[/dim]")
    else:
        console.print("[red]Failed to change password. Current password may be incorrect.[/red]")
        sys.exit(1)


@cli.command()
@click.argument('csv_path')
@click.pass_context
def import_csv(ctx, csv_path: str):
    """Import addresses from a CSV file. CSV must have columns: Account, Coin, Chain, Address, Notes"""
    manager, password = get_manager_with_session(ctx)
    
    if not password:
        console.print("[red]No active session. Please unlock the vault first: key_manager unlock[/red]")
        sys.exit(1)
    
    added, skipped, errors = manager.import_csv(csv_path, password)
    
    console.print(f"[green]✓ Imported {added} addresses[/green]")
    if skipped > 0:
        console.print(f"[yellow]  Skipped {skipped} rows (empty addresses or missing data)[/yellow]")
    if errors:
        console.print(f"[red]  Errors:[/red]")
        for err in errors:
            console.print(f"[red]    {err}[/red]")


@cli.command()
def gen_password():
    """Generate a secure random password."""
    password = generate_secure_password()
    console.print(Panel(
        f"[green]{password}[/green]",
        title="Secure Password",
        subtitle="Copy this password and store it safely"
    ))


@cli.command()
@click.pass_context
def status(ctx):
    """Show vault status and statistics."""
    manager, password = get_manager_with_session(ctx)
    
    vault_exists = manager.data_file.exists()
    session_active = manager.is_session_active()
    
    info_table = Table(title="Vault Status", show_header=False)
    info_table.add_column("Property", style="cyan")
    info_table.add_column("Value", style="green")
    
    info_table.add_row("Vault Location", str(manager.data_file))
    info_table.add_row("Vault Exists", "✓ Yes" if vault_exists else "✗ No")
    info_table.add_row("Session Active", "✓ Yes" if session_active else "✗ No")
    
    if vault_exists:
        file_size = manager.data_file.stat().st_size
        info_table.add_row("File Size", f"{file_size:,} bytes")
    
    if password:
        accounts = manager.address_db.get("accounts", {})
        addr_count = sum(len(acc.get("addresses", [])) for acc in accounts.values())
        info_table.add_row("Accounts Stored", str(len(accounts)))
        info_table.add_row("Total Addresses", str(addr_count))
        info_table.add_row("Mnemonics Stored", str(len(manager.address_db.get("mnemonics", {}))))
    else:
        info_table.add_row("Accounts", "Unlock to view")
    
    console.print(info_table)


if __name__ == '__main__':
    cli()